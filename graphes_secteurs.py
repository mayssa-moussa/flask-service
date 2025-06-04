import pandas as pd
import plotly.graph_objs as go
from openpyxl import load_workbook
import os
import plotly.io as pio


EXCLUDED_INDICATORS = ["Taux d'accessibilité", "Taux de communications réussies"]

def extract_taux_secteur_data(filepath, operators, selected_delegation=None, selected_sector=None):
    print(f"Chargement du fichier : {filepath}")
    wb = load_workbook(filepath, data_only=True)
    ws = wb["Taux_secteur"]

    data = {op: {} for op in operators}
    found_data = False

    # 🔍 Détection des blocs opérateurs via les cellules fusionnées
    operator_blocks = []
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row == merged_range.max_row:  # fusion horizontale
            header_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            if header_cell.value in operators:
                operator_blocks.append({
                    'operator': header_cell.value,
                    'start_row': merged_range.min_row,
                    'start_col': merged_range.min_col,
                    'end_col': merged_range.max_col
                })

    # Trier les blocs par ordre d'apparition dans le fichier
    operator_blocks.sort(key=lambda x: x['start_row'])

    for block in operator_blocks:
        current_operator = block['operator']
        start_col = block['start_col']
        end_col = block['end_col']
        row = block['start_row'] + 1  # Ligne après l'en-tête de l'opérateur

        print(f"Traitement de l'opérateur {current_operator} à partir de la ligne {row}")

        # 🔽 Parcours des blocs de délégations pour l'opérateur courant
        while row <= ws.max_row:
            # Vérifier si la ligne est vide
            if all(ws.cell(row, col).value in [None, "", " "] for col in range(1, end_col + 1)):
                row += 1
                continue

            delegation = ws.cell(row, 1).value
            if not delegation:
                row += 1
                continue

            if selected_delegation and delegation != selected_delegation:
                # Sauter ce bloc
                while row <= ws.max_row and any(ws.cell(row, col).value not in [None, "", " "] for col in range(1, end_col + 1)):
                    row += 1
                continue

            # Récupérer les noms des secteurs
            sectors = []
            for col in range(start_col, end_col + 1):
                secteur = ws.cell(row, col).value
                if secteur and str(secteur).strip().lower() != "total":
                    if selected_sector and str(secteur).strip().lower() != selected_sector.strip().lower():
                        continue
                    sectors.append(f"{delegation} - {secteur}")

            # Lire les indicateurs
            indicator_row = row + 1
            while indicator_row <= ws.max_row:
                cell_val = ws.cell(indicator_row, 1).value
                if not cell_val:
                    break  # Fin du bloc

                indicator = str(cell_val).strip()
                if indicator in EXCLUDED_INDICATORS:
                    indicator_row += 1
                    continue

                values = {}
                for idx, col in enumerate(range(start_col, end_col + 1)):
                    if idx >= len(sectors):
                        break
                    val = ws.cell(indicator_row, col).value
                    values[sectors[idx]] = val if val not in ["--", "NaN", None] else 0

                if indicator not in data[current_operator]:
                    data[current_operator][indicator] = {}
                data[current_operator][indicator].update(values)
                found_data = True
                indicator_row += 1

            row = indicator_row + 1  # Passer à la prochaine délégation

    if not found_data:
        raise ValueError("Aucune donnée correspondante trouvée")

    return data
def plot_interactive_sector_charts(data, selected_sector=None):
    indicators = [ind for ind in next(iter(data.values())).keys() if ind not in EXCLUDED_INDICATORS]
    
    percent_indicators = ['Taux', 'Couverture']
    colors = {
        "Ooredoo TN": "#FF0000",
        "Orange TN": "#FFA500",
        "Tunisie Telecom": "#0000FF"
    }

    graph_blocks = []

    for indicator in indicators:
        fig = go.Figure()
        all_sectors = sorted({s for op in data for s in data[op].get(indicator, {})})
        
        if selected_sector:
            all_sectors = [s for s in all_sectors if selected_sector in s]

        sector_names = [s.split(" - ")[-1] for s in all_sectors]
        is_mos = 'Mos' in indicator

        for operator in data:
            y = []
            for secteur in all_sectors:
                val = data[operator][indicator].get(secteur, 0)
                y.append(val if val not in [None, "NaN", "--"] else 0)

            fig.add_trace(go.Bar(
                x=sector_names,
                y=y,
                name=operator,
                marker_color=colors[operator],
                hovertemplate="<b>%{x}</b><br>%{y:.2%}<extra></extra>" if not is_mos else "<b>%{x}</b><br>%{y:.2f}<extra></extra>",
                text=[f"<b>{v:.2%}</b>" if not is_mos else f"<b>{v:.2f}</b>" for v in y],

                textposition="outside",
                width=0.1
            ))


        fig.update_layout(
            showlegend=False,
            uirevision='persist',
            height=500,
            width=550,
            barmode='group',
            title=dict(
                text=indicator, 
                font=dict(size=20, color='rgb(50, 100, 200)', family='Segoe UI, sans-serif'),
                x=0.5
            ),
            xaxis_title="",
            yaxis_title="Valeur de l'indicateur",
            template="plotly_white",
            xaxis_tickangle=0,
            margin=dict(t=60, b=120),
            yaxis=dict(
                range=[0, 1.2] if not is_mos else [0, 6],
                dtick=0.2 if not is_mos else 1,
                automargin=True,
                gridcolor="lightgrey",
                zeroline=False,
                tickformat=".2%" if not is_mos else ".2f"
            ),
            font=dict(family="Segoe UI, sans-serif", size=14),
        )

        html_block = pio.to_html(
            fig,
            full_html=False,
            include_plotlyjs=False,
            config={"displayModeBar": False}
        )

        # Ajout du bouton "Visualiser sur la carte" uniquement pour Couverture 2G et Couverture 3G
        if indicator in ["Couverture 2G", "Couverture 3G"]:
            button_html = f"""
                <div style="display: flex; justify-content: flex-end; margin-top: 20px;">
                    <button onclick="visualizeOnMap('{indicator}','{secteur}')" 
                        style="
                            padding: 10px 18px;
                            font-size: 14px;
                            background-color: #007BFF;
                            color: white;
                            border: none;
                            border-radius: 6px;
                            cursor: pointer;
                            transition: background-color 0.3s ease, transform 0.2s ease;
                        "
                        onmouseover="this.style.backgroundColor='#0056b3'; this.style.transform='scale(1.05)'"
                        onmouseout="this.style.backgroundColor='#007BFF'; this.style.transform='scale(1)'"
                    >
                        Visualiser sur la carte
                    </button>
                </div>
            """
            html_block += button_html

        graph_blocks.append((indicator, html_block))

    html_content = """
    <html>
    <head>
        <style>
            .graph-row {
                display: grid;
                grid-template-columns: repeat(auto-fill, minmax(500px, 1fr));
                gap: 25px;
                margin: 20px 0;
                width: 100%;
            }

            .graph-section {
                background: white;
                padding: 20px;
                border-radius: 12px;
                box-shadow: 0 2px 8px rgba(0, 0, 0, 0.4);
                transition: transform 0.2s ease;
                overflow: hidden;
                position: relative;
            }

            .graph-section:hover {
                transform: translateY(-3px);
                box-shadow: 0 4px 12px rgba(0, 0, 0, 0.15);
            }

            .plot-container {
                width: 100% !important;
                height: 500px !important;
            }

            .js-plotly-plot {
                height: 100% !important;
                width: 100% !important;
            }

            .modebar-container {
                display: none !important;
            }

            @media (max-width: 768px) {
                .graph-row {
                    grid-template-columns: 1fr;
                    gap: 15px;
                }
                
                .graph-section {
                    padding: 15px;
                    margin: 0 10px;
                }
                
                .plot-container {
                    height: 400px !important;
                }
            }
</style>
    </head>
    <body>
    """

    for i in range(0, len(graph_blocks), 2):
        html_content += '<div class="graph-row">'
        for j in range(i, min(i + 2, len(graph_blocks))):
            html_content += f'<div class="graph-section">{graph_blocks[j][1]}</div>'
        html_content += '</div>'

    html_content += """
    </body>
    </html>
    """

    return html_content

def generate_interactive_graphs_secteur(filepath, selected_operators, selected_delegation=None, selected_sector=None):
    try:
        print("📊 Lecture du fichier Excel...")
        data = extract_taux_secteur_data(
            filepath, 
            selected_operators, 
            selected_delegation, 
            selected_sector
        )
        
        if not any(data.values()):
            return '<div style="color:red">Aucune donnée disponible pour ces paramètres</div>'

        print("📈 Génération des graphiques secteurs...")
        return plot_interactive_sector_charts(data, selected_sector)
    
    except Exception as e:
        return f'<div style="color:red">Erreur: {str(e)}</div>'
