import pandas as pd
import plotly.graph_objs as go
from openpyxl import load_workbook
import os
import plotly.io as pio

def extract_taux_delegation_data(filepath, operators, selected_delegation=None):
    print(f"Chargement du fichier : {filepath}")
    wb = load_workbook(filepath, data_only=True)
    ws = wb["Taux_délégation"]

    data = {}
    row = 1
    found_data = False  # Nouveau flag pour détecter les données
    
    while row <= ws.max_row:
        cell = ws.cell(row=row, column=2)
        if cell.value in operators:
            operator = cell.value
            data[operator] = {}
            delegations = []
            col = 2
            
            # Collecter les délégations
            while True:
                delegation_cell = ws.cell(row=row + 1, column=col)
                if not delegation_cell.value:
                    break
                delegations.append(delegation_cell.value)
                col += 1

            # Filtrer si délégation spécifique
            if selected_delegation and selected_delegation not in delegations:
                row += 1
                continue

            # Extraire les données
            ind_row = row + 2
            while ind_row <= ws.max_row and ws.cell(ind_row, 1).value:
                indicator = ws.cell(ind_row, 1).value
                if indicator in ["Taux d'accessibilité", "Taux de communications réussies"]:
                    ind_row += 1
                    continue

                data[operator][indicator] = {}
                for i, delegation in enumerate(delegations):
                    if selected_delegation and delegation != selected_delegation:
                        continue
                    
                    val = ws.cell(ind_row, i + 2).value
                    data[operator][indicator][delegation] = val if val not in ["--", "NaN", None] else 0
                    found_data = True  # Données trouvées

                ind_row += 1
            row = ind_row + 2
        else:
            row += 1

    if not found_data:
        raise ValueError("Aucune donnée correspondante trouvée")
    
    return data

def plot_interactive_charts(data, output_file="plotly_all_graphs.html"):
    indicators = list(next(iter(data.values())).keys())
    percent_indicators = ['Taux', 'Couverture 2G', 'Couverture 3G']
    mos_indicators = ['Mos Moyen']
    
    colors = {
        "Ooredoo TN": "#FF0000",
        "Orange TN": "#FFA500",
        "Tunisie Telecom": "#0000FF"
    }

    graph_blocks = []

    for indicator in indicators:
        fig = go.Figure()
        
        # Récupérer toutes les délégations sauf "Total"
        all_delegations = sorted({d for operator_data in data.values() for d in operator_data[indicator].keys() if d != "Total"})
        is_mos = any(mos in indicator for mos in mos_indicators)

        for operator in data:
            y, text_labels = [], []
            for d in all_delegations:
                val = data[operator][indicator].get(d)
                if val in [None, "", "NaN"]:
                    val = 0
                if any(p in indicator for p in percent_indicators):
                    display_val = round(val, 4) if val is not None else 0
                    label = f"{display_val * 100:.2f}%"
                else:
                    display_val = round(val, 2) if val is not None else 0
                    label = f"{display_val:.2f}"
                y.append(display_val)
                text_labels.append(label)

            fig.add_trace(go.Bar(
                x=all_delegations,
                y=y,
                name=operator,
                marker_color=colors.get(operator, None),
                customdata=text_labels,
                hovertemplate="%{x}<br>%{customdata}<extra></extra>",
                text=[f"<b>{v:.2%}</b>" if not is_mos else f"<b>{v:.2f}</b>" for v in y],
                textposition="outside",
                textfont=dict(
                    family="Segoe UI, sans-serif",
                    size=13,
                    color="black"
                ),
                width=0.1
            ))

        
        fig.update_layout(
            showlegend=False,
            uirevision='persist',
            height=500,
            width=550,
            barmode='group',
            title=dict(
                text=indicator, 
                font=dict(size=20, color='rgb(50, 100, 200)', family='Segoe UI, sans-serif'),
                x=0.5
            ),
            xaxis_title="Délégation",
            yaxis_title="Valeur de l'indicateur",
            template="plotly_white",
            xaxis_tickangle=0,
            margin=dict(t=60, b=120),
            yaxis=dict(
                range=[0, 1.2] if not is_mos else [0, 6],
                dtick=0.2 if not is_mos else 1,
                automargin=True,
                gridcolor="lightgrey",
                zeroline=False
            ),
            font=dict(family="Segoe UI, sans-serif", size=14),
        )

        if not is_mos:
            fig.update_yaxes(tickformat=".2%", automargin=True)

        # Génération HTML sans barre d'outils
        html_block = pio.to_html(
            fig,
            full_html=False,
            include_plotlyjs=False,
            config={"displayModeBar": False}
        )
        # Ajout du bouton "Visualiser sur la carte" uniquement pour Couverture 2G et Couverture 3G
        # Ajout du bouton "Visualiser sur la carte" uniquement pour Couverture 2G et Couverture 3G
        if indicator in ["Couverture 2G", "Couverture 3G"]:
            button_html = f"""
                <div style="display: flex; justify-content: flex-end; margin-top: 20px;">
                    <button onclick="visualizeOnMap('{indicator}', null)" 
                        style="
                            padding: 10px 18px;
                            font-size: 14px;
                            background-color: #007BFF;
                            color: white;
                            border: none;
                            border-radius: 6px;
                            cursor: pointer;
                            transition: background-color 0.3s ease, transform 0.2s ease;
                        "
                        onmouseover="this.style.backgroundColor='#0056b3'; this.style.transform='scale(1.05)'"
                        onmouseout="this.style.backgroundColor='#007BFF'; this.style.transform='scale(1)'"
                    >
                        Visualiser sur la carte
                    </button>
                </div>
            """
            html_block += button_html
        graph_blocks.append((indicator, html_block))

    # HTML complet avec CSS + chargement unique de Plotly.js
    # HTML complet avec CSS + chargement unique de Plotly.js
    html_content = """
    <html>
    <head>
        
        <style>
            .graph-row {
                display: grid;
                grid-template-columns: repeat(auto-fill, minmax(500px, 1fr));
                gap: 25px;
                margin: 20px 0;
                width: 100%;
            }

            .graph-section {
                background: white;
                padding: 20px;
                border-radius: 12px;
                box-shadow: 0 2px 8px rgba(0, 0, 0, 0.4);
                transition: transform 0.2s ease;
                overflow: hidden;
                position: relative;
            }

            .graph-section:hover {
                transform: translateY(-3px);
                box-shadow: 0 4px 12px rgba(0, 0, 0, 0.15);
            }

            .plot-container {
                width: 100% !important;
                height: 500px !important;
            }

            .js-plotly-plot {
                height: 100% !important;
                width: 100% !important;
            }

            .modebar-container {
                display: none !important;
            }

            @media (max-width: 768px) {
                .graph-row {
                    grid-template-columns: 1fr;
                    gap: 15px;
                }
                
                .graph-section {
                    padding: 15px;
                    margin: 0 10px;
                }
                
                .plot-container {
                    height: 400px !important;
                }
            }
</style>
    </head>
    <body>
    """

    for i in range(0, len(graph_blocks), 2):
        html_content += '<div class="graph-row">'
        for j in range(i, min(i + 2, len(graph_blocks))):
            html_content += f'<div class="graph-section">{graph_blocks[j][1]}</div>'
        html_content += '</div>'

    html_content += """
    </body>
    </html>
    """

    return html_content

def generate_interactive_graphs_délégation(filepath, selected_operators, selected_delegation=None):
    try:
        print("📊 Lecture du fichier Excel...")
        data = extract_taux_delegation_data(filepath, selected_operators, selected_delegation)
        
        if not data:
            return '<div class="alert alert-warning">Aucune donnée trouvée pour cette combinaison</div>'

        print("📈 Génération des graphiques interactifs...")
        html = plot_interactive_charts(data)
        print("✅ Tous les graphes sont prêts.")
        return html
        
    except ValueError as e:
        return f'<div class="alert alert-danger">{str(e)}</div>'
#generate_interactive_graphs_délégation(
#filepath=r"C:\Users\ASUS\Desktop\découpage\Projet2\NABEUL\Statistiques Gouvernorat\Autres Indicateurs.xlsx",
#selected_operators=["Ooredoo TN", "Orange TN", "Tunisie Telecom"],
#selected_delegation="El Haouaria"
#)